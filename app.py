import openpyxl
import os

def main():
    # Insert the file of the spreadsheet you want to change here
    fileName = 'November2019Transactions.xlsx'

    # Make sure to go into the Excel sheet and change the sheet name to transactions prior to running the program
    # Sheet name you would like to change
    sheetName = 'transactions'
    # Load the workbook
    wb = openpyxl.load_workbook(fileName)
    
    # Access the sheet in the workbook
    transactionsWorksheet = wb[sheetName]

    # If E1 is tags, then delete column 5
    delete_tags_column(transactionsWorksheet)
    wb.save(fileName)


    fixDoublePayments(transactionsWorksheet)
    wb.save(fileName)
    
    organizeExpenses(transactionsWorksheet)


def delete_tags_column(transactionsWorksheet):
    tagsValue = transactionsWorksheet['E1'].value
    if (tagsValue == 'Tags'):
        transactionsWorksheet.delete_cols(5)

def organizeExpenses(transactionsWorksheet):
    # Maximum row of the sheet
    maxRow = transactionsWorksheet.max_row

    # Declare empty dictionary to store category as key  and total expense as value
    transactionCategories = {}
    
    # Iterate over each row to obtain category and transition
    for i in range(2, maxRow + 1):
        category = transactionsWorksheet.cell(row=i, column=4).value
        transaction = transactionsWorksheet.cell(row = i, column=5).value

        # If category exists, add the transaction to the total expense
        if category in transactionCategories:
            transactionCategories[category] += transaction
        # If category does not, make a new one and set it to the first transaction that appears
        elif category not in transactionCategories:
            transactionCategories[category] = transaction
    
    for key in transactionCategories:
        print(key, '$', round(transactionCategories[key], 2))
    
def fixDoublePayments(incomeExpensesSheet):
    maxRow = incomeExpensesSheet.max_row

    for i in range(2, maxRow + 1):
        if (str(incomeExpensesSheet.cell(row=i, column= 4).value) == "Credit Card Payments") or (str(incomeExpensesSheet.cell(row=i, column= 3).value) == "Transfer" and str(incomeExpensesSheet.cell(row=i, column= 2).value) == "Trs Plan 3 - Self") or (incomeExpensesSheet.cell(row=i, column= 3).value == "Reinvestment Fidelity 500 Index Fund"):
            cellValue = ('E' + str(i))
            incomeExpensesSheet[cellValue] = 0
            transactionValue = 0

if __name__ == "__main__":
    main()

